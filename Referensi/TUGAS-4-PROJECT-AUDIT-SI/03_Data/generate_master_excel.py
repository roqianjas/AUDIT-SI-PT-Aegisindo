"""
Generate Master Data Excel — Audit SI PT. Murni Solusindo Nusantara
Berisi: Kuesioner, Tabulasi Jawaban 7 Responden, Perhitungan Capability Level, GAP Analysis
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, "Master_Data_Audit_SI.xlsx")

wb = Workbook()

# === STYLES ===
font_title = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
font_sub = Font(name='Calibri', size=11, bold=True, color='1B3A5C')
font_normal = Font(name='Calibri', size=11)
font_bold = Font(name='Calibri', size=11, bold=True)
font_result = Font(name='Calibri', size=12, bold=True, color='0096D6')

fill_header = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
fill_accent = PatternFill(start_color='0096D6', end_color='0096D6', fill_type='solid')
fill_light = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
fill_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
fill_yellow = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
fill_red = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
fill_result = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

def style_cell(ws, row, col, font=font_normal, fill=None, alignment=align_center):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = thin_border
    return cell


# === DATA ===
domains = {
    'APO12': {
        'name': 'Managed Risk (Pengelolaan Risiko)',
        'questions': {
            'Level 1': [
                'Risiko TI yang terkait dengan platform digital telah diidentifikasi secara berkala',
                'Pengumpulan data mengenai risiko TI dilakukan secara efektif',
                'Dokumentasi risiko TI dipelihara dan diperbarui secara teratur',
            ],
            'Level 2': [
                'Terdapat perencanaan untuk pengelolaan risiko TI',
                'Risiko TI dipantau dan dilaporkan kepada pihak terkait secara berkala',
                'Dilakukan evaluasi terhadap pengelolaan risiko secara rutin',
            ],
            'Level 3': [
                'Terdapat prosedur standar yang terdokumentasi untuk mengelola risiko TI',
                'Prosedur pengelolaan risiko diterapkan secara konsisten di seluruh divisi',
                'Dilakukan audit internal terhadap pengelolaan risiko secara periodik',
            ],
        },
        'responses': {
            'Level 1': [
                [3, 4, 3, 4, 3, 3, 2, 3],
                [3, 3, 3, 4, 3, 3, 3, 2],
                [3, 3, 3, 3, 3, 3, 3, 2],
            ],
            'Level 2': [
                [3, 3, 3, 3, 3, 2, 2, 2],
                [3, 3, 3, 3, 2, 2, 2, 2],
                [3, 3, 3, 3, 2, 2, 2, 2],
            ],
            'Level 3': [
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 2, 2, 2, 2, 2, 2, 2],
            ],
        },
    },
    'APO13': {
        'name': 'Managed Security (Pengelolaan Keamanan)',
        'questions': {
            'Level 1': [
                'Kebijakan keamanan informasi telah ditetapkan dan dikomunikasikan',
                'Mekanisme perlindungan terhadap akses tidak sah telah diterapkan',
                'Sosialisasi mengenai keamanan informasi sudah dilakukan kepada karyawan',
            ],
            'Level 2': [
                'Terdapat perencanaan untuk pengelolaan keamanan informasi',
                'Insiden keamanan informasi dipantau dan ditindaklanjuti',
                'Evaluasi terhadap kebijakan keamanan dilakukan secara berkala',
            ],
            'Level 3': [
                'Terdapat prosedur standar keamanan informasi yang terdokumentasi',
                'Prosedur keamanan diterapkan secara konsisten di seluruh platform digital',
                'Dilakukan pelatihan keamanan informasi secara rutin',
            ],
        },
        'responses': {
            'Level 1': [
                [4, 4, 3, 4, 3, 3, 3, 3],
                [4, 3, 3, 4, 3, 3, 3, 3],
                [3, 3, 3, 3, 3, 3, 3, 2],
            ],
            'Level 2': [
                [3, 3, 3, 3, 3, 3, 2, 2],
                [3, 3, 3, 3, 3, 2, 2, 2],
                [3, 3, 3, 3, 2, 3, 2, 2],
            ],
            'Level 3': [
                [3, 3, 2, 3, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
            ],
        },
    },
    'BAI06': {
        'name': 'Managed IT Changes (Pengelolaan Perubahan TI)',
        'questions': {
            'Level 1': [
                'Setiap perubahan pada sistem TI dicatat dan didokumentasikan',
                'Terdapat proses persetujuan sebelum implementasi perubahan',
                'Evaluasi dampak perubahan dilakukan sebelum implementasi',
            ],
            'Level 2': [
                'Perubahan TI direncanakan dan dijadwalkan secara terstruktur',
                'Terdapat komunikasi mengenai perubahan kepada pihak terkait',
                'Evaluasi pasca-implementasi perubahan dilakukan secara rutin',
            ],
            'Level 3': [
                'Terdapat prosedur standar change management yang terdokumentasi',
                'Prosedur change management diterapkan konsisten untuk semua perubahan',
                'Dilakukan review berkala terhadap efektivitas prosedur change management',
            ],
        },
        'responses': {
            'Level 1': [
                [4, 4, 3, 4, 3, 3, 2, 3],
                [4, 3, 3, 3, 3, 3, 3, 2],
                [3, 3, 3, 3, 3, 3, 3, 2],
            ],
            'Level 2': [
                [3, 3, 3, 3, 3, 3, 2, 2],
                [3, 3, 3, 3, 2, 3, 2, 2],
                [3, 3, 3, 3, 2, 2, 2, 2],
            ],
            'Level 3': [
                [3, 3, 2, 3, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
            ],
        },
    },
    'DSS01': {
        'name': 'Managed Operations (Pengelolaan Operasional)',
        'questions': {
            'Level 1': [
                'Pemantauan terhadap kinerja platform digital dilakukan secara rutin',
                'Terdapat prosedur penanganan insiden operasional',
                'Jadwal pemeliharaan infrastruktur TI sudah ditetapkan',
            ],
            'Level 2': [
                'Operasional TI direncanakan dan dikelola secara terstruktur',
                'Kinerja operasional TI dipantau dan dilaporkan secara berkala',
                'Evaluasi terhadap pengelolaan operasional dilakukan rutin',
            ],
            'Level 3': [
                'Terdapat SOP operasional TI yang terdokumentasi secara formal',
                'SOP operasional diterapkan secara konsisten di seluruh divisi',
                'Terdapat mekanisme feedback dari pengguna terhadap layanan TI',
            ],
        },
        'responses': {
            'Level 1': [
                [4, 4, 4, 4, 3, 3, 3, 3],
                [4, 4, 3, 3, 3, 3, 3, 3],
                [4, 3, 3, 3, 3, 3, 3, 2],
            ],
            'Level 2': [
                [3, 3, 3, 3, 3, 3, 3, 2],
                [3, 3, 3, 3, 3, 3, 2, 2],
                [3, 3, 3, 3, 3, 2, 2, 2],
            ],
            'Level 3': [
                [3, 3, 3, 2, 2, 2, 2, 2],
                [3, 3, 3, 2, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
            ],
        },
    },
    'DSS05': {
        'name': 'Managed Security Services (Pengelolaan Layanan Keamanan)',
        'questions': {
            'Level 1': [
                'Perlindungan terhadap malware dan ancaman siber telah diterapkan',
                'Pengelolaan hak akses pengguna (access control) sudah dilaksanakan',
                'Prosedur pencadangan dan pemulihan data sudah berjalan',
            ],
            'Level 2': [
                'Layanan keamanan TI direncanakan dan dikelola secara terstruktur',
                'Insiden keamanan TI dipantau dan ditindaklanjuti secara berkala',
                'Evaluasi efektivitas layanan keamanan dilakukan rutin',
            ],
            'Level 3': [
                'Terdapat prosedur standar pengelolaan layanan keamanan yang terdokumentasi',
                'Standar keamanan diterapkan konsisten di seluruh platform digital',
                'Audit keamanan internal dilaksanakan secara periodik',
            ],
        },
        'responses': {
            'Level 1': [
                [4, 4, 4, 4, 3, 3, 3, 3],
                [4, 4, 3, 4, 3, 3, 3, 3],
                [4, 3, 3, 3, 3, 3, 3, 2],
            ],
            'Level 2': [
                [3, 3, 3, 3, 3, 2, 2, 2],
                [3, 3, 3, 3, 2, 2, 2, 2],
                [3, 3, 3, 3, 2, 2, 2, 2],
            ],
            'Level 3': [
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 3, 2, 2, 2, 2, 2, 2],
                [3, 2, 2, 2, 2, 2, 2, 2],
            ],
        },
    },
}

respondents = [
    ('R1', 'Kepala Divisi IT'),
    ('R2', 'Senior Web Developer'),
    ('R3', 'Junior Web Developer'),
    ('R4', 'Content & SEO Specialist'),
    ('R5', 'Staff IT Support'),
    ('R6', 'Digital Marketing'),
    ('R7', 'Staff Finance'),
    ('R8', 'Staff HR'),
]


# ============================================================
# SHEET 1: PROFIL RESPONDEN
# ============================================================
ws = wb.active
ws.title = 'Profil Responden'
ws.sheet_properties.tabColor = '1B3A5C'

ws.merge_cells('A1:F1')
ws['A1'] = 'PROFIL RESPONDEN KUESIONER'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1B3A5C')
ws['A1'].alignment = align_center

ws.merge_cells('A2:F2')
ws['A2'] = 'Audit Sistem Informasi — PT. Murni Solusindo Nusantara'
ws['A2'].font = Font(name='Calibri', size=11)
ws['A2'].alignment = align_center

headers = ['No', 'Kode', 'Jabatan', 'Divisi', 'Lama Bekerja', 'Pendidikan']
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, 6)

data = [
    (1, 'R1', 'Kepala Divisi IT', 'IT', '8 tahun', 'S1 Teknik Informatika'),
    (2, 'R2', 'Senior Web Developer', 'Digital Marketing (Web Dev & SEO)', '5 tahun', 'S1 Sistem Informasi'),
    (3, 'R3', 'Junior Web Developer', 'Digital Marketing (Web Dev & SEO)', '2 tahun', 'S1 Sistem Informasi'),
    (4, 'R4', 'Content & SEO Specialist', 'Digital Marketing (Web Dev & SEO)', '4 tahun', 'S1 Teknik Informatika'),
    (5, 'R5', 'Staff IT Support', 'IT', '3 tahun', 'D3 Teknik Komputer'),
    (6, 'R6', 'Digital Marketing Specialist', 'Digital Marketing', '3 tahun', 'S1 Ilmu Komunikasi'),
    (7, 'R7', 'Staff Finance', 'Finance', '4 tahun', 'S1 Akuntansi'),
    (8, 'R8', 'Staff HR', 'Human Resources', '3 tahun', 'S1 Manajemen SDM'),
]
for i, row in enumerate(data):
    for j, val in enumerate(row):
        cell = ws.cell(row=5+i, column=j+1, value=val)
        cell.font = font_normal
        cell.alignment = align_center if j < 2 else align_left
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 25


# ============================================================
# SHEET 2-6: TABULASI PER DOMAIN
# ============================================================
for domain_code, domain_data in domains.items():
    ws = wb.create_sheet(title=f'Tabulasi {domain_code}')
    ws.sheet_properties.tabColor = '0096D6'

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = f'TABULASI DATA KUESIONER — {domain_code}: {domain_data["name"]}'
    ws['A1'].font = Font(name='Calibri', size=13, bold=True, color='1B3A5C')
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:M2')
    ws['A2'] = 'PT. Murni Solusindo Nusantara | Skala: STS=1, TS=2, R=3, S=4, SS=5'
    ws['A2'].font = Font(name='Calibri', size=10, color='666666')
    ws['A2'].alignment = align_center

    # Headers
    headers = ['No', 'Pernyataan'] + [r[0] for r in respondents] + ['Total', 'Mean']
    row = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    row = 5
    q_num = 1
    for level_name, questions in domain_data['questions'].items():
        # Level header
        ws.merge_cells(f'A{row}:M{row}')
        ws.cell(row=row, column=1, value=level_name)
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=11, bold=True, color='0096D6')
        ws.cell(row=row, column=1).fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        ws.cell(row=row, column=1).alignment = align_left
        for c in range(1, len(headers)+1):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        responses = domain_data['responses'][level_name]
        for qi, question in enumerate(questions):
            ws.cell(row=row, column=1, value=q_num).font = font_normal
            ws.cell(row=row, column=1).alignment = align_center
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=question).font = font_normal
            ws.cell(row=row, column=2).alignment = align_left
            ws.cell(row=row, column=2).border = thin_border

            resp = responses[qi]
            total = sum(resp)
            mean = total / len(resp)

            for ri, val in enumerate(resp):
                cell = ws.cell(row=row, column=3+ri, value=val)
                cell.font = font_normal
                cell.alignment = align_center
                cell.border = thin_border
                if qi % 2 == 1:
                    cell.fill = fill_light

            # Total
            cell = ws.cell(row=row, column=11, value=total)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = thin_border

            # Mean
            cell = ws.cell(row=row, column=12, value=round(mean, 2))
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = thin_border
            cell.number_format = '0.00'

            q_num += 1
            row += 1

        # Level summary
        level_responses = responses
        all_means = [sum(r)/len(r) for r in level_responses]
        avg_mean = sum(all_means) / len(all_means)
        pct = (avg_mean / 5) * 100

        if pct > 85:
            rating = 'F (Fully Achieved)'
            rfill = fill_green
        elif pct > 50:
            rating = 'L (Largely Achieved)'
            rfill = fill_yellow
        elif pct > 15:
            rating = 'P (Partially Achieved)'
            rfill = fill_red
        else:
            rating = 'N (Not Achieved)'
            rfill = fill_red

        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value=f'Rata-rata {level_name}')
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).alignment = align_left
        for c in range(1, 13):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = rfill

        ws.cell(row=row, column=12, value=round(avg_mean, 2))
        ws.cell(row=row, column=12).font = font_result
        ws.cell(row=row, column=12).alignment = align_center
        ws.cell(row=row, column=12).number_format = '0.00'

        row += 1

        # Percentage + Rating
        ws.merge_cells(f'A{row}:I{row}')
        ws.cell(row=row, column=1, value=f'% Ketercapaian: {pct:.2f}% → Rating: {rating}')
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10, bold=True, color='0096D6')
        ws.cell(row=row, column=1).alignment = align_left
        for c in range(1, 13):
            ws.cell(row=row, column=c).border = thin_border
        row += 1
        row += 1  # blank row

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 55
    for c in range(3, 11):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.column_dimensions['K'].width = 8
    ws.column_dimensions['L'].width = 8


# ============================================================
# SHEET 7: RINGKASAN CAPABILITY LEVEL
# ============================================================
ws = wb.create_sheet(title='Ringkasan Capability Level')
ws.sheet_properties.tabColor = '00BC8C'

ws.merge_cells('A1:H1')
ws['A1'] = 'RINGKASAN PERHITUNGAN CAPABILITY LEVEL'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1B3A5C')
ws['A1'].alignment = align_center

ws.merge_cells('A2:H2')
ws['A2'] = 'Framework COBIT 2019 | PT. Murni Solusindo Nusantara'
ws['A2'].font = Font(name='Calibri', size=11, color='666666')
ws['A2'].alignment = align_center

headers = ['No', 'Domain', 'Level 1 (%)', 'Rating L1', 'Level 2 (%)', 'Rating L2',
           'Level 3 (%)', 'Rating L3', 'Capability Level']
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, len(headers))

cap_data = [
    (1, 'APO12 — Managed Risk', 60.83, 'L', 51.67, 'L', 45.00, 'P', 2),
    (2, 'APO13 — Managed Security', 65.00, 'L', 54.17, 'L', 46.67, 'P', 2),
    (3, 'BAI06 — Managed IT Changes', 62.50, 'L', 53.33, 'L', 46.67, 'P', 2),
    (4, 'DSS01 — Managed Operations', 66.67, 'L', 55.83, 'L', 48.33, 'P', 2),
    (5, 'DSS05 — Managed Security Services', 67.50, 'L', 51.67, 'L', 45.00, 'P', 2),
]

for i, row_data in enumerate(cap_data):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light

        # Color ratings
        if col in [4, 6] and val == 'L':
            cell.fill = fill_green
            cell.font = Font(name='Calibri', size=11, bold=True, color='2E7D32')
        elif col == 8 and val == 'P':
            cell.fill = fill_red
            cell.font = Font(name='Calibri', size=11, bold=True, color='E53935')

        # Percentage format
        if col in [3, 5, 7]:
            cell.number_format = '0.00'

        # Capability level
        if col == 9:
            cell.font = font_result
            cell.fill = fill_result

    # Domain name left-aligned
    ws.cell(row=row, column=2).alignment = align_left

# Average row
row = 10
ws.merge_cells(f'A{row}:B{row}')
ws.cell(row=row, column=1, value='RATA-RATA')
ws.cell(row=row, column=1).font = font_bold
ws.cell(row=row, column=1).alignment = align_center
for c in range(1, 10):
    ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=c).fill = fill_result
ws.cell(row=row, column=3, value=64.50).font = font_bold
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=5, value=53.33).font = font_bold
ws.cell(row=row, column=5).alignment = align_center
ws.cell(row=row, column=7, value=46.33).font = font_bold
ws.cell(row=row, column=7).alignment = align_center
ws.cell(row=row, column=9, value=2).font = font_result
ws.cell(row=row, column=9).alignment = align_center

# GAP Analysis section
row = 12
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1, value='GAP ANALYSIS')
ws.cell(row=row, column=1).font = Font(name='Calibri', size=13, bold=True, color='1B3A5C')

row = 13
gap_headers = ['No', 'Domain', 'Capability Level (As-is)', 'Target Level (To-be)', 'GAP']
for col, h in enumerate(gap_headers, 1):
    ws.cell(row=row, column=col, value=h)
style_header_row(ws, row, len(gap_headers))

gap_data = [
    (1, 'APO12 — Managed Risk', 2, 4, 2),
    (2, 'APO13 — Managed Security', 2, 4, 2),
    (3, 'BAI06 — Managed IT Changes', 2, 4, 2),
    (4, 'DSS01 — Managed Operations', 2, 4, 2),
    (5, 'DSS05 — Managed Security Services', 2, 4, 2),
]
for i, row_data in enumerate(gap_data):
    row = 14 + i
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        if col == 5:
            cell.fill = fill_red
            cell.font = Font(name='Calibri', size=11, bold=True, color='E53935')
    ws.cell(row=row, column=2).alignment = align_left

# Average GAP
row = 19
ws.merge_cells(f'A{row}:B{row}')
ws.cell(row=row, column=1, value='RATA-RATA GAP')
ws.cell(row=row, column=1).font = font_bold
ws.cell(row=row, column=1).alignment = align_center
for c in range(1, 6):
    ws.cell(row=row, column=c).border = thin_border
    ws.cell(row=row, column=c).fill = fill_result
ws.cell(row=row, column=3, value=2).font = font_bold
ws.cell(row=row, column=3).alignment = align_center
ws.cell(row=row, column=4, value=4).font = font_bold
ws.cell(row=row, column=4).alignment = align_center
ws.cell(row=row, column=5, value=2).font = font_result
ws.cell(row=row, column=5).alignment = align_center

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 35
for c in range(3, 10):
    ws.column_dimensions[get_column_letter(c)].width = 18


# ============================================================
# SHEET 8: KETERANGAN SKALA
# ============================================================
ws = wb.create_sheet(title='Keterangan Skala')
ws.sheet_properties.tabColor = 'FFD600'

ws.merge_cells('A1:D1')
ws['A1'] = 'KETERANGAN SKALA DAN RATING'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1B3A5C')
ws['A1'].alignment = align_center

# Skala Likert
ws.merge_cells('A3:D3')
ws['A3'] = 'Skala Likert Kuesioner'
ws['A3'].font = font_sub

headers = ['Skor', 'Keterangan', 'Deskripsi']
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, 3)

likert = [
    (1, 'STS', 'Sangat Tidak Setuju'),
    (2, 'TS', 'Tidak Setuju'),
    (3, 'R', 'Ragu-ragu / Netral'),
    (4, 'S', 'Setuju'),
    (5, 'SS', 'Sangat Setuju'),
]
for i, (skor, ket, desc) in enumerate(likert):
    for j, val in enumerate([skor, ket, desc], 1):
        cell = ws.cell(row=5+i, column=j, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border

# Rating Scale
row = 12
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'Rating Scale COBIT 2019'
ws[f'A{row}'].font = font_sub

headers = ['Notasi', 'Deskripsi', 'Persentase Ketercapaian']
for col, h in enumerate(headers, 1):
    ws.cell(row=row+1, column=col, value=h)
style_header_row(ws, row+1, 3)

ratings = [
    ('N', 'Not Achieved', '0 – 15%'),
    ('P', 'Partially Achieved', '> 15% – 50%'),
    ('L', 'Largely Achieved', '> 50% – 85%'),
    ('F', 'Fully Achieved', '> 85% – 100%'),
]
fills_rating = [fill_red, PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'), fill_yellow, fill_green]
for i, (notasi, desc, pct) in enumerate(ratings):
    for j, val in enumerate([notasi, desc, pct], 1):
        cell = ws.cell(row=row+2+i, column=j, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = fills_rating[i]

# Capability Level
row = 19
ws.merge_cells(f'A{row}:D{row}')
ws[f'A{row}'] = 'Capability Level COBIT 2019'
ws[f'A{row}'].font = font_sub

headers = ['Level', 'Nama', 'Deskripsi']
for col, h in enumerate(headers, 1):
    ws.cell(row=row+1, column=col, value=h)
style_header_row(ws, row+1, 3)

cap_levels = [
    (0, 'Incomplete', 'Kurang memiliki kemampuan dasar'),
    (1, 'Performed', 'Proses kurang lebih mencapai tujuan, bersifat initial/intuitive'),
    (2, 'Managed', 'Proses mencapai tujuan melalui kegiatan dasar yang lengkap'),
    (3, 'Established', 'Proses terdefinisi baik, kinerja diukur kualitatif'),
    (4, 'Predictable', 'Proses terdefinisi baik, kinerja diukur kuantitatif'),
    (5, 'Optimizing', 'Proses terdefinisi, diukur, dan perbaikan berkelanjutan'),
]
for i, (level, name, desc) in enumerate(cap_levels):
    for j, val in enumerate([level, name, desc], 1):
        cell = ws.cell(row=row+2+i, column=j, value=val)
        cell.font = font_normal
        cell.alignment = align_center if j < 3 else align_left
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 20


# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT)
print(f"✅ Master Data Excel berhasil disimpan: {OUTPUT}")
print(f"📊 Total sheet: {len(wb.sheetnames)}")
print(f"   Sheet: {', '.join(wb.sheetnames)}")
print(f"📋 Isi: Profil Responden, Tabulasi 5 Domain, Ringkasan Capability Level, Keterangan Skala")
