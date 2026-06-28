"""
Generate Master Data Excel — Audit SI PT Aegisindo Mitra Sejati
Berisi: Profil Responden, Tabulasi Jawaban 6 Responden per Domain,
Ringkasan Capability Level, GAP Analysis, dan Keterangan Skala.
Seluruh perhitungan dihitung dinamis dari data jawaban (single source of truth).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT = os.path.join(BASE_DIR, "Master_Data_Audit_SI.xlsx")

wb = Workbook()

# === COLORS (safety palette) ===
HDR = 'F26522'      # safety orange
ACCENT = 'C65512'   # darker orange
RESULT_TXT = 'B26A00'

font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
font_sub = Font(name='Calibri', size=11, bold=True, color=ACCENT)
font_normal = Font(name='Calibri', size=11)
font_bold = Font(name='Calibri', size=11, bold=True)
font_result = Font(name='Calibri', size=12, bold=True, color=ACCENT)

fill_header = PatternFill(start_color=HDR, end_color=HDR, fill_type='solid')
fill_light = PatternFill(start_color='FDEDE3', end_color='FDEDE3', fill_type='solid')
fill_green = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
fill_yellow = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
fill_red = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
fill_result = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border


def rating_info(pct):
    if pct > 85:
        return 'F (Fully Achieved)', fill_green
    if pct > 50:
        return 'L (Largely Achieved)', fill_yellow
    if pct > 15:
        return 'P (Partially Achieved)', fill_red
    return 'N (Not Achieved)', fill_red


def capability_from(levels_pct):
    """Capability = level tertinggi yang mencapai L (>50%) secara berurutan dari Level 1."""
    cap = 0
    for lvl in sorted(levels_pct.keys()):
        if levels_pct[lvl] > 50:
            cap = lvl
        else:
            break
    return cap


# === DATA: jawaban 6 responden (R1..R6). Single source of truth ===
respondents = ['R1', 'R2', 'R3', 'R4', 'R5', 'R6']

domains = {
    'APO14': {
        'name': 'Managed Data (Pengelolaan Data)',
        'target': 3,
        'questions': {
            1: ['Pendataan data inventori, pelanggan, penjualan',
                'Pemutakhiran data produk & stok saat transaksi',
                'Penyimpanan data pelanggan & riwayat transaksi'],
            2: ['Perencanaan data & penanggung jawab',
                'Pemantauan kualitas & konsistensi data antar saluran',
                'Evaluasi akurasi & duplikasi data'],
            3: ['Prosedur standar pengelolaan & pemutakhiran data',
                'Kebijakan klasifikasi & kepemilikan data',
                'Integrasi data tersentralisasi antar saluran'],
        },
        'responses': {
            1: [[4, 3, 3, 4, 3, 2], [4, 3, 3, 3, 3, 2], [3, 3, 3, 3, 3, 2]],
            2: [[3, 3, 2, 3, 2, 2], [3, 2, 2, 2, 2, 2], [3, 2, 2, 2, 2, 2]],
            3: [[2, 2, 2, 2, 2, 1], [2, 2, 2, 2, 2, 1], [2, 2, 2, 2, 1, 1]],
        },
    },
    'BAI09': {
        'name': 'Managed Assets (Pengelolaan Aset)',
        'target': 3,
        'questions': {
            1: ['Pencatatan aset persediaan barang safety',
                'Pencatatan barang masuk & keluar gudang',
                'Mengetahui jumlah & lokasi persediaan'],
            2: ['Perencanaan persediaan (min. stok, reorder point)',
                'Pemantauan & pelaporan posisi stok berkala',
                'Stock opname catatan vs fisik secara rutin'],
            3: ['Prosedur standar pengelolaan persediaan & aset',
                'Catatan stok terintegrasi gudang-toko-marketplace',
                'Siklus hidup aset terdokumentasi'],
        },
        'responses': {
            1: [[4, 4, 3, 4, 3, 3], [4, 4, 3, 4, 3, 3], [4, 3, 3, 3, 3, 3]],
            2: [[3, 3, 3, 3, 3, 2], [3, 3, 3, 3, 2, 2], [3, 3, 2, 3, 2, 2]],
            3: [[3, 2, 2, 3, 2, 2], [3, 2, 2, 2, 2, 2], [2, 2, 2, 2, 2, 2]],
        },
    },
    'DSS01': {
        'name': 'Managed Operations (Pengelolaan Operasional)',
        'target': 4,
        'questions': {
            1: ['Pemenuhan pesanan dari berbagai saluran rutin',
                'Prosedur penanganan kendala operasional',
                'Aktivitas operasional harian teratur'],
            2: ['Perencanaan & penjadwalan operasional',
                'Pemantauan & pelaporan kinerja operasional',
                'Evaluasi proses operasional'],
            3: ['SOP operasional terdokumentasi semua saluran',
                'Konsistensi penerapan SOP oleh staf',
                'Mekanisme umpan balik pelanggan'],
        },
        'responses': {
            1: [[4, 4, 4, 4, 3, 3], [4, 4, 3, 4, 3, 3], [4, 3, 3, 4, 3, 3]],
            2: [[4, 3, 3, 3, 3, 2], [3, 3, 3, 3, 3, 2], [3, 3, 3, 3, 2, 2]],
            3: [[3, 3, 2, 3, 2, 2], [3, 2, 2, 3, 2, 2], [3, 2, 2, 2, 2, 2]],
        },
    },
    'DSS04': {
        'name': 'Managed Continuity (Pengelolaan Keberlangsungan)',
        'target': 3,
        'questions': {
            1: ['Kesadaran risiko gangguan saluran penjualan',
                'Upaya backup data penting',
                'Saluran alternatif saat satu saluran terganggu'],
            2: ['Perencanaan business continuity terstruktur',
                'Backup data terjadwal & dipantau',
                'Evaluasi kesiapan menghadapi gangguan'],
            3: ['Dokumen BCP/DRP terdokumentasi formal',
                'Uji coba prosedur pemulihan berkala',
                'Penetapan target RTO & RPO'],
        },
        'responses': {
            1: [[4, 3, 3, 3, 3, 2], [3, 3, 3, 3, 3, 2], [3, 3, 3, 3, 3, 2]],
            2: [[3, 2, 2, 3, 2, 2], [3, 2, 2, 2, 2, 2], [2, 2, 2, 2, 2, 2]],
            3: [[2, 2, 2, 2, 2, 1], [2, 2, 2, 2, 1, 1], [2, 2, 1, 2, 1, 1]],
        },
    },
    'DSS05': {
        'name': 'Managed Security Services (Pengelolaan Layanan Keamanan)',
        'target': 3,
        'questions': {
            1: ['Perlindungan perangkat dari malware & virus',
                'Pengaturan hak akses akun & data',
                'Perlindungan kata sandi akun penting'],
            2: ['Perencanaan pengelolaan keamanan akun & data',
                'Pemantauan akses & pelaporan insiden',
                'Evaluasi efektivitas pengamanan'],
            3: ['Prosedur standar keamanan informasi terdokumentasi',
                'Tinjauan hak akses berkala',
                'Pemeriksaan keamanan internal berkala'],
        },
        'responses': {
            1: [[4, 4, 3, 4, 3, 3], [4, 3, 3, 4, 3, 3], [3, 3, 3, 3, 3, 3]],
            2: [[3, 3, 3, 3, 3, 2], [3, 3, 2, 3, 2, 2], [3, 3, 2, 3, 2, 2]],
            3: [[3, 2, 2, 2, 2, 2], [2, 2, 2, 2, 2, 2], [2, 2, 2, 2, 2, 2]],
        },
    },
}

LEVEL_NAMES = {1: 'Level 1 (Performed)', 2: 'Level 2 (Managed)', 3: 'Level 3 (Established)'}


# ============================================================
# SHEET 1: PROFIL RESPONDEN
# ============================================================
ws = wb.active
ws.title = 'Profil Responden'
ws.sheet_properties.tabColor = HDR

ws.merge_cells('A1:F1')
ws['A1'] = 'PROFIL RESPONDEN KUESIONER'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT)
ws['A1'].alignment = align_center
ws.merge_cells('A2:F2')
ws['A2'] = 'Audit Sistem Informasi — PT Aegisindo Mitra Sejati'
ws['A2'].font = Font(name='Calibri', size=11)
ws['A2'].alignment = align_center

headers = ['No', 'Kode', 'Jabatan', 'Divisi', 'Lama Bekerja', 'Pendidikan']
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, 6)

data = [
    (1, 'R1', 'Manajer Operasional', 'Operasional', '7 tahun', 'S1 Manajemen'),
    (2, 'R2', 'Staff Procurement & Gudang', 'Pengadaan & Gudang', '5 tahun', 'S1 Teknik Industri'),
    (3, 'R3', 'Staff Penjualan Online (Marketplace)', 'Penjualan & Marketing', '4 tahun', 'S1 Manajemen'),
    (4, 'R4', 'Staff IT / Admin Sistem', 'IT / Administrasi', '3 tahun', 'S1 Sistem Informasi'),
    (5, 'R5', 'Staff Keuangan & Administrasi', 'Keuangan & Administrasi', '5 tahun', 'S1 Akuntansi'),
    (6, 'R6', 'Staff Customer Service', 'Customer Service & After Sales', '2 tahun', 'D3 Administrasi Bisnis'),
]
for i, row in enumerate(data):
    for j, val in enumerate(row):
        cell = ws.cell(row=5+i, column=j+1, value=val)
        cell.font = font_normal
        cell.alignment = align_center if j < 2 else align_left
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 28
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 25


# ============================================================
# SHEET 2-6: TABULASI PER DOMAIN (dihitung dinamis)
# ============================================================
summary = []  # kumpulan hasil untuk sheet ringkasan

for code, d in domains.items():
    ws = wb.create_sheet(title=f'Tabulasi {code}')
    ws.sheet_properties.tabColor = ACCENT

    last_col = 2 + len(respondents) + 2  # No, Pernyataan, R1..R6, Total, Mean
    last_letter = get_column_letter(last_col)

    ws.merge_cells(f'A1:{last_letter}1')
    ws['A1'] = f'TABULASI DATA KUESIONER — {code}: {d["name"]}'
    ws['A1'].font = Font(name='Calibri', size=13, bold=True, color=ACCENT)
    ws['A1'].alignment = align_center
    ws.merge_cells(f'A2:{last_letter}2')
    ws['A2'] = 'PT Aegisindo Mitra Sejati | Skala: STS=1, TS=2, R=3, S=4, SS=5'
    ws['A2'].font = Font(name='Calibri', size=10, color='666666')
    ws['A2'].alignment = align_center

    hdrs = ['No', 'Pernyataan'] + respondents + ['Total', 'Mean']
    for col, h in enumerate(hdrs, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(hdrs))

    mean_col = len(hdrs)
    total_col = len(hdrs) - 1
    row = 5
    q_num = 1
    levels_pct = {}

    for lvl in (1, 2, 3):
        ws.merge_cells(f'A{row}:{last_letter}{row}')
        ws.cell(row=row, column=1, value=LEVEL_NAMES[lvl])
        c = ws.cell(row=row, column=1)
        c.font = Font(name='Calibri', size=11, bold=True, color=ACCENT)
        c.fill = fill_result
        c.alignment = align_left
        for cc in range(1, len(hdrs)+1):
            ws.cell(row=row, column=cc).border = thin_border
        row += 1

        level_means = []
        for qi, question in enumerate(d['questions'][lvl]):
            resp = d['responses'][lvl][qi]
            total = sum(resp)
            mean = total / len(resp)
            level_means.append(mean)

            ws.cell(row=row, column=1, value=q_num).font = font_normal
            ws.cell(row=row, column=1).alignment = align_center
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=question).font = font_normal
            ws.cell(row=row, column=2).alignment = align_left
            ws.cell(row=row, column=2).border = thin_border
            for ri, val in enumerate(resp):
                cell = ws.cell(row=row, column=3+ri, value=val)
                cell.font = font_normal
                cell.alignment = align_center
                cell.border = thin_border
                if qi % 2 == 1:
                    cell.fill = fill_light
            cell = ws.cell(row=row, column=total_col, value=total)
            cell.font = font_bold; cell.alignment = align_center; cell.border = thin_border
            cell = ws.cell(row=row, column=mean_col, value=round(mean, 2))
            cell.font = font_bold; cell.alignment = align_center; cell.border = thin_border
            cell.number_format = '0.00'
            q_num += 1
            row += 1

        avg_mean = sum(level_means) / len(level_means)
        pct = (avg_mean / 5) * 100
        levels_pct[lvl] = pct
        rating, rfill = rating_info(pct)

        ws.merge_cells(f'A{row}:{get_column_letter(total_col-1)}{row}')
        ws.cell(row=row, column=1, value=f'Rata-rata {LEVEL_NAMES[lvl]}')
        ws.cell(row=row, column=1).font = font_bold
        ws.cell(row=row, column=1).alignment = align_left
        for cc in range(1, len(hdrs)+1):
            ws.cell(row=row, column=cc).border = thin_border
            ws.cell(row=row, column=cc).fill = rfill
        ws.cell(row=row, column=mean_col, value=round(avg_mean, 2))
        ws.cell(row=row, column=mean_col).font = font_result
        ws.cell(row=row, column=mean_col).alignment = align_center
        ws.cell(row=row, column=mean_col).number_format = '0.00'
        row += 1

        ws.merge_cells(f'A{row}:{get_column_letter(total_col-1)}{row}')
        ws.cell(row=row, column=1, value=f'% Ketercapaian: {pct:.2f}% → Rating: {rating}')
        ws.cell(row=row, column=1).font = Font(name='Calibri', size=10, bold=True, color=ACCENT)
        ws.cell(row=row, column=1).alignment = align_left
        for cc in range(1, len(hdrs)+1):
            ws.cell(row=row, column=cc).border = thin_border
        row += 2

    cap = capability_from(levels_pct)
    summary.append({'code': code, 'name': d['name'], 'l1': levels_pct[1],
                    'l2': levels_pct[2], 'l3': levels_pct[3], 'cap': cap, 'target': d['target']})

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    for cc in range(3, 3 + len(respondents)):
        ws.column_dimensions[get_column_letter(cc)].width = 6
    ws.column_dimensions[get_column_letter(total_col)].width = 8
    ws.column_dimensions[get_column_letter(mean_col)].width = 8


# ============================================================
# SHEET 7: RINGKASAN CAPABILITY LEVEL + GAP
# ============================================================
ws = wb.create_sheet(title='Ringkasan Capability Level')
ws.sheet_properties.tabColor = 'FFB300'

ws.merge_cells('A1:I1')
ws['A1'] = 'RINGKASAN PERHITUNGAN CAPABILITY LEVEL'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT)
ws['A1'].alignment = align_center
ws.merge_cells('A2:I2')
ws['A2'] = 'Framework COBIT 2019 | PT Aegisindo Mitra Sejati'
ws['A2'].font = Font(name='Calibri', size=11, color='666666')
ws['A2'].alignment = align_center

headers = ['No', 'Domain', 'Level 1 (%)', 'Rating L1', 'Level 2 (%)', 'Rating L2',
           'Level 3 (%)', 'Rating L3', 'Capability Level']
for col, h in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, len(headers))

def short_rating(pct):
    if pct > 85: return 'F'
    if pct > 50: return 'L'
    if pct > 15: return 'P'
    return 'N'

for i, s in enumerate(summary):
    row = 5 + i
    vals = [i+1, f"{s['code']} — {s['name'].split('(')[0].strip()}",
            round(s['l1'], 2), short_rating(s['l1']),
            round(s['l2'], 2), short_rating(s['l2']),
            round(s['l3'], 2), short_rating(s['l3']), s['cap']]
    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light
        if col in (4, 6, 8):
            if val == 'L' or val == 'F':
                cell.fill = fill_green
                cell.font = Font(name='Calibri', size=11, bold=True, color='2E7D32')
            else:
                cell.fill = fill_red
                cell.font = Font(name='Calibri', size=11, bold=True, color='C62828')
        if col in (3, 5, 7):
            cell.number_format = '0.00'
        if col == 9:
            cell.font = font_result
            cell.fill = fill_result
    ws.cell(row=row, column=2).alignment = align_left

# average row
n = len(summary)
avg_row = 5 + n
ws.merge_cells(f'A{avg_row}:B{avg_row}')
ws.cell(row=avg_row, column=1, value='RATA-RATA')
ws.cell(row=avg_row, column=1).font = font_bold
ws.cell(row=avg_row, column=1).alignment = align_center
for c in range(1, 10):
    ws.cell(row=avg_row, column=c).border = thin_border
    ws.cell(row=avg_row, column=c).fill = fill_result
ws.cell(row=avg_row, column=3, value=round(sum(s['l1'] for s in summary)/n, 2)).font = font_bold
ws.cell(row=avg_row, column=5, value=round(sum(s['l2'] for s in summary)/n, 2)).font = font_bold
ws.cell(row=avg_row, column=7, value=round(sum(s['l3'] for s in summary)/n, 2)).font = font_bold
ws.cell(row=avg_row, column=9, value=round(sum(s['cap'] for s in summary)/n, 1)).font = font_result
for c in (3, 5, 7, 9):
    ws.cell(row=avg_row, column=c).alignment = align_center
    ws.cell(row=avg_row, column=c).number_format = '0.00'

# GAP section
gap_title_row = avg_row + 2
ws.merge_cells(f'A{gap_title_row}:I{gap_title_row}')
ws.cell(row=gap_title_row, column=1, value='GAP ANALYSIS')
ws.cell(row=gap_title_row, column=1).font = Font(name='Calibri', size=13, bold=True, color=ACCENT)

gap_hdr_row = gap_title_row + 1
gap_headers = ['No', 'Domain', 'Capability Level (As-is)', 'Target Level (To-be)', 'GAP']
for col, h in enumerate(gap_headers, 1):
    ws.cell(row=gap_hdr_row, column=col, value=h)
style_header_row(ws, gap_hdr_row, len(gap_headers))

for i, s in enumerate(summary):
    row = gap_hdr_row + 1 + i
    gap = s['target'] - s['cap']
    vals = [i+1, f"{s['code']} — {s['name'].split('(')[0].strip()}", s['cap'], s['target'], gap]
    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border
        if col == 5:
            cell.fill = fill_red
            cell.font = Font(name='Calibri', size=11, bold=True, color='C62828')
    ws.cell(row=row, column=2).alignment = align_left

gap_avg_row = gap_hdr_row + 1 + n
ws.merge_cells(f'A{gap_avg_row}:B{gap_avg_row}')
ws.cell(row=gap_avg_row, column=1, value='RATA-RATA')
ws.cell(row=gap_avg_row, column=1).font = font_bold
ws.cell(row=gap_avg_row, column=1).alignment = align_center
for c in range(1, 6):
    ws.cell(row=gap_avg_row, column=c).border = thin_border
    ws.cell(row=gap_avg_row, column=c).fill = fill_result
ws.cell(row=gap_avg_row, column=3, value=round(sum(s['cap'] for s in summary)/n, 1)).font = font_bold
ws.cell(row=gap_avg_row, column=4, value=round(sum(s['target'] for s in summary)/n, 1)).font = font_bold
ws.cell(row=gap_avg_row, column=5, value=round(sum(s['target']-s['cap'] for s in summary)/n, 1)).font = font_result
for c in (3, 4, 5):
    ws.cell(row=gap_avg_row, column=c).alignment = align_center
    ws.cell(row=gap_avg_row, column=c).number_format = '0.0'

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 38
for c in range(3, 10):
    ws.column_dimensions[get_column_letter(c)].width = 18


# ============================================================
# SHEET 8: KETERANGAN SKALA
# ============================================================
ws = wb.create_sheet(title='Keterangan Skala')
ws.sheet_properties.tabColor = '5B7080'

ws.merge_cells('A1:C1')
ws['A1'] = 'KETERANGAN SKALA DAN RATING'
ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=ACCENT)
ws['A1'].alignment = align_center

ws.merge_cells('A3:C3')
ws['A3'] = 'Skala Likert Kuesioner'
ws['A3'].font = font_sub
for col, h in enumerate(['Skor', 'Singkatan', 'Keterangan'], 1):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, 3)
likert = [(1, 'STS', 'Sangat Tidak Setuju'), (2, 'TS', 'Tidak Setuju'),
          (3, 'R', 'Ragu-ragu / Netral'), (4, 'S', 'Setuju'), (5, 'SS', 'Sangat Setuju')]
for i, vals in enumerate(likert):
    for j, val in enumerate(vals, 1):
        cell = ws.cell(row=5+i, column=j, value=val)
        cell.font = font_normal; cell.alignment = align_center; cell.border = thin_border

row = 12
ws.merge_cells(f'A{row}:C{row}')
ws[f'A{row}'] = 'Rating Scale COBIT 2019'
ws[f'A{row}'].font = font_sub
for col, h in enumerate(['Notasi', 'Deskripsi', 'Persentase'], 1):
    ws.cell(row=row+1, column=col, value=h)
style_header_row(ws, row+1, 3)
ratings = [('N', 'Not Achieved', '0 – 15%'), ('P', 'Partially Achieved', '> 15% – 50%'),
           ('L', 'Largely Achieved', '> 50% – 85%'), ('F', 'Fully Achieved', '> 85% – 100%')]
fills_rating = [fill_red, PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'), fill_yellow, fill_green]
for i, vals in enumerate(ratings):
    for j, val in enumerate(vals, 1):
        cell = ws.cell(row=row+2+i, column=j, value=val)
        cell.font = font_normal; cell.alignment = align_center; cell.border = thin_border
        cell.fill = fills_rating[i]

row = 19
ws.merge_cells(f'A{row}:C{row}')
ws[f'A{row}'] = 'Capability Level COBIT 2019'
ws[f'A{row}'].font = font_sub
for col, h in enumerate(['Level', 'Nama', 'Deskripsi'], 1):
    ws.cell(row=row+1, column=col, value=h)
style_header_row(ws, row+1, 3)
cap_levels = [
    (0, 'Incomplete', 'Proses tidak dilaksanakan atau gagal mencapai tujuan'),
    (1, 'Performed', 'Proses dilaksanakan, tujuan tercapai, masih intuitif'),
    (2, 'Managed', 'Proses direncanakan, dipantau, dilaporkan; kegiatan dasar lengkap'),
    (3, 'Established', 'Proses terdefinisi baik via prosedur standar, diukur kualitatif'),
    (4, 'Predictable', 'Proses diukur kuantitatif sehingga kinerja dapat diprediksi'),
    (5, 'Optimizing', 'Proses terus diperbaiki secara berkelanjutan'),
]
for i, vals in enumerate(cap_levels):
    for j, val in enumerate(vals, 1):
        cell = ws.cell(row=row+2+i, column=j, value=val)
        cell.font = font_normal
        cell.alignment = align_center if j < 3 else align_left
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = fill_light

ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 58


wb.save(OUTPUT)
print(f"OK Master Data Excel: {OUTPUT}")
print(f"Total sheet: {len(wb.sheetnames)} -> {', '.join(wb.sheetnames)}")
for s in summary:
    print(f"  {s['code']}: L1={s['l1']:.2f}% L2={s['l2']:.2f}% L3={s['l3']:.2f}% -> Cap Level {s['cap']} (target {s['target']}, GAP {s['target']-s['cap']})")
